import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import random
import json
from openpyxl import Workbook, load_workbook


class HealthyLifestylePlanner(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("Healthy Lifestyle Planner")
        self.geometry("463x700")

        self.sports_list = []
        self.food_list = []
        self.activities_list = []

        self.selected_items = {"Sports": [], "Food Plans": [], "Activities": []}

        self.create_widgets()

    def create_widgets(self):
        self.import_button = tk.Button(self, text="Import Excel File", command=self.import_file)
        self.import_button.grid(row=0, column=0, columnspan=2, pady=10)

        self.combo = ttk.Combobox(self, values=["Sports", "Food Plans", "Activities"])
        self.combo.set("Sports")
        self.combo.grid(row=1, column=0, padx=10, pady=10)

        self.listbox = tk.Listbox(self, selectmode=tk.SINGLE)
        self.listbox.grid(row=2, column=0, rowspan=4, padx=10, pady=10, sticky="nsew")
        self.load_items("Sports")

        self.selected_listbox = tk.Listbox(self)
        self.selected_listbox.grid(row=2, column=1, rowspan=4, padx=10, pady=10, sticky="nsew")

        self.plan_text = tk.Text(self, height=8, width=40)
        self.plan_text.grid(row=9, column=0, columnspan=2, padx=10, pady=10)

        self.add_button = tk.Button(self, text="Add", command=self.add_item)
        self.add_button.grid(row=6, column=0, padx=10, pady=5, sticky="ew")

        self.delete_button = tk.Button(self, text="Delete", command=self.delete_item)
        self.delete_button.grid(row=7, column=0, padx=10, pady=5, sticky="ew")

        self.random_button = tk.Button(self, text="Randomize", command=self.randomize)
        self.random_button.grid(row=8, column=0, padx=10, pady=5, sticky="ew")

        self.week_label = tk.Label(self, text="Week Number:")
        self.week_label.grid(row=10, column=0, padx=10, pady=10, sticky="w")
        self.week_entry = tk.Entry(self)
        self.week_entry.grid(row=10, column=1, padx=10, pady=10)

        self.file_type_label = tk.Label(self, text="Select File Type:")
        self.file_type_label.grid(row=11, column=0, padx=10, pady=10, sticky="w")

        self.file_type_combo = ttk.Combobox(self, values=["TXT", "JSON", "XLSX"])
        self.file_type_combo.set("TXT")
        self.file_type_combo.grid(row=11, column=1, padx=10, pady=10, sticky="ew")

        self.export_button = tk.Button(self, text="Export", command=self.export_data)
        self.export_button.grid(row=12, column=0, columnspan=2, padx=10, pady=10)

        self.combo.bind("<<ComboboxSelected>>", self.update_listbox)

    def load_items(self, category):
        self.listbox.delete(0, tk.END)

        if category == "Sports":
            items = self.sports_list
        elif category == "Food Plans":
            items = self.food_list
        else:
            items = self.activities_list

        for item in items:
            self.listbox.insert(tk.END, item)

    def add_item(self):
        selected_item = self.listbox.get(tk.ACTIVE)
        current_category = self.combo.get()

        if selected_item and len(self.selected_items[current_category]) < 7:
            if selected_item not in self.selected_items[current_category]:
                self.selected_items[current_category].append(selected_item)
                self.selected_listbox.insert(tk.END, selected_item)
                self.update_plan_display()
            else:
                messagebox.showerror("Error", "Item already selected!")
        else:
            messagebox.showerror("Error", "You can only select 7 items.")

    def delete_item(self):
        selected_item = self.selected_listbox.get(tk.ACTIVE)
        current_category = self.combo.get()

        if selected_item:
            self.selected_items[current_category].remove(selected_item)
            self.selected_listbox.delete(tk.ACTIVE)
            self.update_plan_display()
        else:
            messagebox.showerror("Error", "Please select an item to delete.")

    def update_listbox(self, event):
        category = self.combo.get()
        self.load_items(category)
        self.selected_listbox.delete(0, tk.END)
        for item in self.selected_items[category]:
            self.selected_listbox.insert(tk.END, item)
        self.update_plan_display()

    def randomize(self):
        self.selected_items = {"Sports": [], "Food Plans": [], "Activities": []}
        self.selected_listbox.delete(0, tk.END)
        self.plan_text.delete(1.0, tk.END)

        random_sports = random.sample(self.sports_list, 7)
        random_foods = random.sample(self.food_list, 7)
        random_activities = random.sample(self.activities_list, 7)

        self.selected_items["Sports"] = random_sports
        self.selected_items["Food Plans"] = random_foods
        self.selected_items["Activities"] = random_activities

        self.selected_listbox.insert(tk.END, *random_sports, *random_foods, *random_activities)
        self.update_plan_display()

    def update_plan_display(self):
        self.plan_text.delete(1.0, tk.END)

        for category, items in self.selected_items.items():
            if items:
                self.plan_text.insert(tk.END, f"{category}:\n")
                for item in items:
                    self.plan_text.insert(tk.END, f"{item}\n")
                self.plan_text.insert(tk.END, "\n")

    def import_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            try:
                wb = load_workbook(file_path)
                sheet = wb.active

                self.sports_list = []
                self.food_list = []
                self.activities_list = []

                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                    self.sports_list.append(row[0].value)
                    self.food_list.append(row[1].value)
                    self.activities_list.append(row[2].value)

                self.load_items("Sports")
                messagebox.showinfo("Success", "Excel file imported successfully!")

            except Exception as e:
                messagebox.showerror("Error", f"Error importing file: {e}")

    def export_data(self):
        week = self.week_entry.get() or "1"
        if not week.isdigit():
            messagebox.showerror("Error", "Invalid week number")
            return

        file_type = self.file_type_combo.get().lower()
        selected_file = filedialog.asksaveasfilename(defaultextension=f".{file_type}", filetypes=[("Text files", "*.txt"), ("JSON files", "*.json"), ("Excel files", "*.xlsx")])

        if selected_file:
            try:
                if file_type == "json":
                    with open(selected_file, "w") as json_file:
                        json.dump(self.selected_items, json_file)
                elif file_type == "xlsx":
                    wb = Workbook()
                    ws = wb.active
                    ws.title = f"Week {week} Plan"
                    ws.append(["Category", "Items"])
                    for category, items in self.selected_items.items():
                        for item in items:
                            ws.append([category, item])
                    wb.save(selected_file)
                elif file_type == "txt":
                    with open(selected_file, "w") as txt_file:
                        for category, items in self.selected_items.items():
                            txt_file.write(f"{category}:\n")
                            txt_file.write("\n".join(items) + "\n")
                else:
                    messagebox.showerror("Error", "Unsupported file type!")
                messagebox.showinfo("Success", "File exported successfully!")
            except Exception as e:
                messagebox.showerror("Error", f"Error exporting file: {e}")


if __name__ == "__main__":
    app = HealthyLifestylePlanner()
    app.mainloop()
